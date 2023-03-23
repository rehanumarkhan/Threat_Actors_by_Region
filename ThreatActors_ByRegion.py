import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font

# Fetch data from MITRE ATT&CK Enterprise Framework API
def fetch_mitre_data():
    url = "https://raw.githubusercontent.com/mitre/cti/master/enterprise-attack/enterprise-attack.json"
    response = requests.get(url)
    return response.json()

# Process fetched data into a DataFrame
def process_data(raw_data):
    tactics = {}
    techniques = []
    threat_actors = []

    # First, process all tactics
    for item in raw_data['objects']:
        if item['type'] == 'x-mitre-tactic':
            tactics[item['x_mitre_shortname']] = item['name']

    # Then, process all techniques
    for item in raw_data['objects']:
        if item['type'] == 'attack-pattern':
            for phase in item['kill_chain_phases']:
                if phase['kill_chain_name'] == 'mitre-attack' and phase['phase_name'] in tactics:
                    technique = {
                        'tactic_id': phase['phase_name'],
                        'tactic': tactics[phase['phase_name']],
                        'technique_id': item['external_references'][0]['external_id'],
                        'technique': item['name']
                    }
                    techniques.append(technique)

        # Process threat actors
        if item['type'] == 'intrusion-set':
            threat_actor = {
                'threat_actor_id': item['external_references'][0]['external_id'],
                'threat_actor': item['name'],
                'aliases': ', '.join(item['aliases']) if 'aliases' in item else '',
                'description': item['description'] if 'description' in item else ""
            }
            threat_actors.append(threat_actor)

    df_techniques = pd.DataFrame(techniques)
    df_threat_actors = pd.DataFrame(threat_actors)

    return df_techniques, df_threat_actors, tactics

def filter_threat_actors_by_region(df_threat_actors, region):
    filtered_threat_actors = df_threat_actors[df_threat_actors['description'].str.contains(region, case=False)]
    return filtered_threat_actors

def save_to_excel(df, sheet_name, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Set header styles
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
        if ws.row_dimensions[ws.max_row].height is None:
            ws.row_dimensions[ws.max_row].height = 45
            for cell in ws[ws.max_row]:
                cell.font = header_font
                cell.alignment = header_alignment

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
                ws.column_dimensions[column].width = max_length + 2

    wb.save(filename)

if __name__ == "__main__":
    raw_data = fetch_mitre_data()
    df_techniques, df_threat_actors, tactics = process_data(raw_data)

    target_region = "Middle East"  # Replace this with the desired region
    filtered_threat_actors = filter_threat_actors_by_region(df_threat_actors, target_region)

    if filtered_threat_actors is not None:
        save_to_excel(filtered_threat_actors, 'filtered_threat_actors', f'{target_region}_ThreatActors.xlsx')
    else:
        print(f"No threat actors found targeting {target_region}")